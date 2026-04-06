"""
운전자 근로시간 분석 대시보드 v1.0
- 운행 단말기 데이터 기반 개인별 실제 근로시간 산출
- 이상값(미조작/과장조작) 자동 탐지 및 보정
- 개인별 근무일수 / 만근일수 산출
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, date
import calendar
import io
import warnings
from openpyxl import load_workbook
warnings.filterwarnings("ignore")

# ============================================================
# 앱 기본 설정
# ============================================================
st.set_page_config(
    page_title="운전자 근로시간 분석",
    page_icon="🚌",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── 브라우저 자동번역 차단 (Chrome / 네이버 웨일 번역 충돌 방지) ──
st.markdown("""
<style>
  .goog-te-banner-frame { display: none !important; }
  body { top: 0 !important; }
</style>
<script>
(function() {
  // 1) 즉시 실행: 페이지 언어 고정 + notranslate 클래스 추가
  var html = document.documentElement;
  html.setAttribute('lang', 'ko');
  html.setAttribute('translate', 'no');
  html.classList.add('notranslate');

  // 2) head에 meta 태그 삽입
  function injectMeta() {
    if (document.head) {
      var m1 = document.createElement('meta');
      m1.name = 'google'; m1.content = 'notranslate';
      document.head.insertBefore(m1, document.head.firstChild);
      var m2 = document.createElement('meta');
      m2.httpEquiv = 'Content-Language'; m2.content = 'ko';
      document.head.insertBefore(m2, document.head.firstChild);
    }
  }
  if (document.head) { injectMeta(); }
  else { document.addEventListener('DOMContentLoaded', injectMeta); }

  // 3) removeChild 패치: 번역기가 React 노드를 잘못 삭제할 때 에러 방지
  var origRemoveChild = Node.prototype.removeChild;
  Node.prototype.removeChild = function(child) {
    if (child && child.parentNode === this) {
      return origRemoveChild.call(this, child);
    }
    // 자식이 아닌 노드 제거 시도 → 무시 (번역기 충돌 방지)
    console.warn('[notranslate] removeChild 무시됨 (번역기 충돌 방지)');
    return child;
  };

  // 4) insertBefore 패치
  var origInsertBefore = Node.prototype.insertBefore;
  Node.prototype.insertBefore = function(newNode, refNode) {
    if (refNode && refNode.parentNode !== this) {
      console.warn('[notranslate] insertBefore 무시됨 (번역기 충돌 방지)');
      return newNode;
    }
    return origInsertBefore.call(this, newNode, refNode);
  };
})();
</script>
""", unsafe_allow_html=True)

# ============================================================
# 한국 공휴일 (공휴일에 관한 규정 / 대체공휴일 포함)
# 2020 ~ 2025
# ============================================================
KOREAN_HOLIDAYS = {
    # ─── 2020 ───
    date(2020, 1, 1): "신정",
    date(2020, 1, 24): "설날 연휴",
    date(2020, 1, 25): "설날",
    date(2020, 1, 26): "설날 연휴",
    date(2020, 1, 27): "설날 대체공휴일",
    date(2020, 3, 1): "삼일절",
    date(2020, 4, 15): "국회의원선거일",
    date(2020, 4, 30): "부처님오신날 대체공휴일",
    date(2020, 5, 1): "근로자의날",
    date(2020, 5, 5): "어린이날",
    date(2020, 6, 6): "현충일",
    date(2020, 8, 15): "광복절",
    date(2020, 8, 17): "광복절 대체공휴일",
    date(2020, 9, 30): "추석 연휴",
    date(2020, 10, 1): "추석",
    date(2020, 10, 2): "추석 연휴",
    date(2020, 10, 3): "개천절",
    date(2020, 10, 9): "한글날",
    date(2020, 12, 25): "성탄절",
    # ─── 2021 ───
    date(2021, 1, 1): "신정",
    date(2021, 2, 11): "설날 연휴",
    date(2021, 2, 12): "설날",
    date(2021, 2, 13): "설날 연휴",
    date(2021, 3, 1): "삼일절",
    date(2021, 5, 1): "근로자의날",
    date(2021, 5, 5): "어린이날",
    date(2021, 5, 19): "부처님오신날",
    date(2021, 6, 6): "현충일",
    date(2021, 8, 15): "광복절",
    date(2021, 8, 16): "광복절 대체공휴일",
    date(2021, 9, 20): "추석 연휴",
    date(2021, 9, 21): "추석",
    date(2021, 9, 22): "추석 연휴",
    date(2021, 10, 3): "개천절",
    date(2021, 10, 4): "추석 대체공휴일",
    date(2021, 10, 9): "한글날",
    date(2021, 10, 11): "한글날 대체공휴일",
    date(2021, 12, 25): "성탄절",
    # ─── 2022 ───
    date(2022, 1, 1): "신정",
    date(2022, 1, 31): "설날 연휴",
    date(2022, 2, 1): "설날",
    date(2022, 2, 2): "설날 연휴",
    date(2022, 3, 1): "삼일절",
    date(2022, 3, 9): "대통령선거일",
    date(2022, 5, 1): "근로자의날",
    date(2022, 5, 5): "어린이날",
    date(2022, 5, 8): "어린이날 대체공휴일",
    date(2022, 5, 9): "부처님오신날",
    date(2022, 6, 1): "지방선거일",
    date(2022, 6, 6): "현충일",
    date(2022, 8, 15): "광복절",
    date(2022, 9, 9): "추석 연휴",
    date(2022, 9, 10): "추석",
    date(2022, 9, 11): "추석 연휴",
    date(2022, 9, 12): "추석 대체공휴일",
    date(2022, 10, 3): "개천절",
    date(2022, 10, 9): "한글날",
    date(2022, 10, 10): "한글날 대체공휴일",
    date(2022, 12, 25): "성탄절",
    date(2022, 12, 26): "성탄절 대체공휴일",
    # ─── 2023 ───
    date(2023, 1, 1): "신정",
    date(2023, 1, 21): "설날 연휴",
    date(2023, 1, 22): "설날",
    date(2023, 1, 23): "설날 연휴",
    date(2023, 1, 24): "설날 대체공휴일",
    date(2023, 3, 1): "삼일절",
    date(2023, 5, 1): "근로자의날",
    date(2023, 5, 5): "어린이날",
    date(2023, 5, 27): "부처님오신날",
    date(2023, 5, 29): "부처님오신날 대체공휴일",
    date(2023, 6, 6): "현충일",
    date(2023, 8, 15): "광복절",
    date(2023, 9, 28): "추석 연휴",
    date(2023, 9, 29): "추석",
    date(2023, 9, 30): "추석 연휴",
    date(2023, 10, 2): "임시공휴일",
    date(2023, 10, 3): "개천절",
    date(2023, 10, 9): "한글날",
    date(2023, 12, 25): "성탄절",
    # ─── 2024 ───
    date(2024, 1, 1): "신정",
    date(2024, 2, 9): "설날 연휴",
    date(2024, 2, 10): "설날",
    date(2024, 2, 11): "설날 연휴",
    date(2024, 2, 12): "설날 대체공휴일",
    date(2024, 3, 1): "삼일절",
    date(2024, 4, 10): "국회의원선거일",
    date(2024, 5, 1): "근로자의날",
    date(2024, 5, 5): "어린이날",
    date(2024, 5, 6): "어린이날 대체공휴일",
    date(2024, 5, 15): "부처님오신날",
    date(2024, 6, 6): "현충일",
    date(2024, 8, 15): "광복절",
    date(2024, 9, 16): "추석 연휴",
    date(2024, 9, 17): "추석",
    date(2024, 9, 18): "추석 연휴",
    date(2024, 10, 3): "개천절",
    date(2024, 10, 9): "한글날",
    date(2024, 12, 25): "성탄절",
    # ─── 2025 ───
    date(2025, 1, 1): "신정",
    date(2025, 1, 28): "설날 연휴",
    date(2025, 1, 29): "설날",
    date(2025, 1, 30): "설날 연휴",
    date(2025, 3, 1): "삼일절",
    date(2025, 3, 3): "삼일절 대체공휴일",
    date(2025, 5, 1): "근로자의날",
    date(2025, 5, 5): "어린이날",
    date(2025, 5, 6): "부처님오신날",
    date(2025, 6, 6): "현충일",
    date(2025, 8, 15): "광복절",
    date(2025, 10, 3): "개천절",
    date(2025, 10, 5): "추석 연휴",
    date(2025, 10, 6): "추석",
    date(2025, 10, 7): "추석 연휴",
    date(2025, 10, 8): "추석 대체공휴일",
    date(2025, 10, 9): "한글날",
    date(2025, 12, 25): "성탄절",
}

WEEKDAY_MAP = {"월": 0, "화": 1, "수": 2, "목": 3, "금": 4, "토": 5, "일": 6}
WEEKDAY_REV = {v: k for k, v in WEEKDAY_MAP.items()}

# ============================================================
# 핵심 처리 함수
# ============================================================

def parse_trip_minutes(start_str, end_str):
    """운행출발~종료 시각(HH:MM:SS)을 분 단위 운행시간으로 변환. 자정 넘김 처리."""
    try:
        s = str(start_str)
        e = str(end_str)
        sh, sm = int(s[0:2]), int(s[3:5])
        eh, em = int(e[0:2]), int(e[3:5])
        s_mins = sh * 60 + sm
        e_mins = eh * 60 + em
        if e_mins < s_mins:         # 자정을 넘기는 운행
            e_mins += 1440
        return e_mins - s_mins
    except Exception:
        return np.nan


def get_start_minutes(start_str):
    """운행출발일시 → 분(0~1439)"""
    try:
        s = str(start_str)
        return int(s[0:2]) * 60 + int(s[3:5])
    except Exception:
        return np.nan


@st.cache_data(show_spinner=False)
def load_operation_files(file_list):
    """운행데이터 xlsx 여러 파일 → 통합 DataFrame"""
    dfs = []
    errors = []
    for f in file_list:
        try:
            f.seek(0)
            df = pd.read_excel(io.BytesIO(f.read()), header=2)
            df = df.dropna(subset=["운행일"])
            dfs.append(df)
        except Exception as ex:
            errors.append(f"{f.name}: {ex}")
    if errors:
        for e in errors:
            st.warning(f"⚠️ 파일 로드 실패 → {e}")
    if not dfs:
        return pd.DataFrame()

    out = pd.concat(dfs, ignore_index=True)
    out.drop_duplicates(inplace=True)

    # 기본 파생 컬럼
    out["운행일"] = out["운행일"].astype(str).str.strip()
    out["운행일_dt"] = pd.to_datetime(out["운행일"], format="%Y%m%d", errors="coerce")
    out["연도"] = out["운행일_dt"].dt.year.astype("Int64")
    out["월"] = out["운행일_dt"].dt.month.astype("Int64")
    out["요일"] = out["운행일_dt"].dt.dayofweek   # 0=월 … 6=일
    out["노선"] = out["노선"].astype(str).str.strip()
    out["운전자"] = out["운전자"].astype(str).str.strip()
    return out


@st.cache_data(show_spinner=False)
def load_route_ref(file):
    """인가현황 파일 → 노선별 인가운행시간 참조 테이블"""
    file.seek(0)
    df = pd.read_excel(io.BytesIO(file.read()), header=2)
    df = df.dropna(subset=["노선"])
    df = df.rename(columns={"운행시간": "인가운행시간", "적용년": "연도"})
    df["노선"] = df["노선"].astype(str).str.strip()
    df["연도"] = pd.to_numeric(df["연도"], errors="coerce").astype("Int64")
    df["인가운행시간"] = pd.to_numeric(df["인가운행시간"], errors="coerce")
    return df[["회사명", "연도", "노선", "인가운행시간"]].drop_duplicates()


@st.cache_data(show_spinner=False)
def process_data(_df_raw, _route_ref, short_min, extra_min):
    """
    메인 처리:
    1) 타임스탬프 기반 운행시간 계산
    2) 노선별 인가운행시간 조인
    3) 이상값 탐지 (단시간 / 장시간)
    4) 이상값 보정 (같은 월·요일·노선·출발±60분 정상 평균)
    """
    df = _df_raw.copy()

    # ── 1. 운행시간 계산 ──
    df["계산운행시간"] = df.apply(
        lambda r: parse_trip_minutes(r["운행출발일시"], r["운행종료일시"]), axis=1
    )
    df["출발분"] = df["운행출발일시"].apply(get_start_minutes)

    # ── 2. 인가운행시간 조인 ──
    if _route_ref is not None and not _route_ref.empty:
        df = df.merge(
            _route_ref[["노선", "연도", "인가운행시간"]].drop_duplicates(),
            on=["노선", "연도"], how="left"
        )
    else:
        df["인가운행시간"] = np.nan

    # ── 3. 이상값 탐지 ──
    df["이상값_상한"] = df["인가운행시간"].fillna(300) + extra_min
    df["이상값여부"] = (
        (df["계산운행시간"] < short_min) |
        (df["계산운행시간"] > df["이상값_상한"])
    )
    df["이상값_유형"] = ""
    df.loc[df["계산운행시간"] < short_min, "이상값_유형"] = "단시간(미조작)"
    df.loc[df["계산운행시간"] > df["이상값_상한"], "이상값_유형"] = "장시간(과장조작)"

    # ── 4. 이상값 보정 ──
    # 정상 레코드로 참조 테이블 구성 (출발 2시간 버킷 × 노선 × 요일 × 연도·월)
    normal = df[~df["이상값여부"]].copy()
    normal["출발_버킷"] = (normal["출발분"] // 120) * 120   # 2시간 단위

    ref = (
        normal.groupby(["연도", "월", "노선", "요일", "출발_버킷"])["계산운행시간"]
        .mean()
        .reset_index()
        .rename(columns={"계산운행시간": "보정기준시간"})
    )
    # fallback: 연도·월·노선 평균
    ref_fallback = (
        normal.groupby(["연도", "월", "노선"])["계산운행시간"]
        .mean()
        .reset_index()
        .rename(columns={"계산운행시간": "보정기준시간_fallback"})
    )

    df["출발_버킷"] = (df["출발분"] // 120) * 120
    df = df.merge(ref, on=["연도", "월", "노선", "요일", "출발_버킷"], how="left")
    df = df.merge(ref_fallback, on=["연도", "월", "노선"], how="left")

    df["보정운행시간"] = df["계산운행시간"].copy().astype(float)
    anom_mask = df["이상값여부"]
    # 1차 대체: 버킷 기반 평균
    df.loc[anom_mask & df["보정기준시간"].notna(), "보정운행시간"] = \
        df.loc[anom_mask & df["보정기준시간"].notna(), "보정기준시간"]
    # 2차 대체: 노선·월 평균
    still_anom = anom_mask & df["보정운행시간"].isna()
    df.loc[still_anom & df["보정기준시간_fallback"].notna(), "보정운행시간"] = \
        df.loc[still_anom & df["보정기준시간_fallback"].notna(), "보정기준시간_fallback"]
    # 3차 대체: 인가운행시간 사용
    still_anom2 = anom_mask & df["보정운행시간"].isna()
    df.loc[still_anom2, "보정운행시간"] = df.loc[still_anom2, "인가운행시간"].fillna(120)

    df["보정운행시간"] = df["보정운행시간"].round(1)
    return df


def build_daily_v2(df_proc):
    """처리된 데이터 → 개인별 일별 집계 (버그 없는 버전)"""
    daily = df_proc.groupby(
        ["운전자", "운행일_dt", "운행일", "연도", "월", "요일"]
    ).agg(
        운행횟수=("보정운행시간", "count"),
        총운행시간_분=("보정운행시간", "sum"),
        이상값포함=("이상값여부", "any"),
    ).reset_index()
    daily["총운행시간_시간"] = (daily["총운행시간_분"] / 60).round(2)
    return daily


def calc_mangeun(year, month, hd1, hd2):
    """
    만근일수 = 달력일수 - 지정휴일수 - 공휴일수(지정휴일 요일 제외)
    hd1, hd2: 지정휴일 요일번호 (0=월 … 6=일)
    """
    _, days = calendar.monthrange(year, month)
    # 해당 월의 지정휴일 날짜 집합
    designated = set()
    for d in range(1, days + 1):
        dt = date(year, month, d)
        if dt.weekday() in (hd1, hd2):
            designated.add(dt)
    # 해당 월의 법정공휴일
    pub_hols = {d for d in KOREAN_HOLIDAYS if d.year == year and d.month == month}
    # 지정휴일과 겹치지 않는 공휴일만 차감
    extra = pub_hols - designated
    mangeun = days - len(designated) - len(extra)
    return {
        "달력일수": days,
        "지정휴일수": len(designated),
        "비중복_공휴일수": len(extra),
        "만근일수": mangeun,
        "공휴일목록": sorted(extra),
    }


def monthly_attendance(df_daily, driver_hol_df, year, month):
    """월별 운전자별 근무일수 + 만근일수 테이블"""
    sub = df_daily[(df_daily["연도"] == year) & (df_daily["월"] == month)]
    wdays = sub.groupby("운전자")["운행일_dt"].nunique().reset_index()
    wdays.columns = ["운전자", "근무일수"]
    time_sum = sub.groupby("운전자")["총운행시간_시간"].sum().reset_index()
    wdays = wdays.merge(time_sum, on="운전자", how="left")

    if driver_hol_df is None or driver_hol_df.empty:
        wdays["만근일수"] = "─"
        wdays["초과/미달"] = "─"
        return wdays

    # 컬럼명 공백/BOM 제거 (CSV 저장 방식에 따라 컬럼명에 공백이 붙을 수 있음)
    hol_df = driver_hol_df.copy()
    hol_df.columns = [c.strip() for c in hol_df.columns]

    # 운전자 컬럼명 자동 탐지 (운전자 / 이름 / name 등 허용)
    driver_col = None
    for cand in ["운전자", "이름", "성명", "name", "driver"]:
        if cand in hol_df.columns:
            driver_col = cand
            break
    if driver_col is None:
        driver_col = hol_df.columns[0]  # 첫 번째 컬럼을 운전자로 간주

    hol_df = hol_df.rename(columns={driver_col: "운전자"})
    hol_df["운전자"] = hol_df["운전자"].astype(str).str.strip()

    rows = []
    for _, drow in hol_df.iterrows():
        driver = drow["운전자"]
        hd1 = int(drow.get("지정휴일1", 5))
        hd2 = int(drow.get("지정휴일2", 6))
        mg = calc_mangeun(year, month, hd1, hd2)
        rows.append({"운전자": driver, **mg})
    mg_df = pd.DataFrame(rows)

    result = wdays.merge(
        mg_df[["운전자", "달력일수", "지정휴일수", "비중복_공휴일수", "만근일수"]],
        on="운전자", how="left"
    )
    result["초과/미달"] = result["근무일수"] - result["만근일수"].fillna(0)
    return result



def calc_night_minutes(start_m, trip_mins):
    """회차별 야간(22:00~06:00) 겹치는 시간(분) 계산"""
    if start_m is None or np.isnan(start_m): return 0.0
    end_m = start_m + trip_mins
    night  = max(0, min(end_m, 360)  - max(start_m, 0))      # 00:00~06:00
    night += max(0, min(end_m, 1440) - max(start_m, 1320))    # 22:00~24:00
    night += max(0, min(end_m, 1800) - max(start_m, 1440))    # 다음날 00:00~06:00
    return float(max(0, night))


def classify_work_hours(df_proc, driver_hol_df):
    """
    처리된 운행 데이터 → 일별 근로시간 6분류
    A 평일8h기준 / B 평일9h월상계 / C 야간 / D 쉬프트단축 / E 휴일8h / F 휴일8h초과
    """
    df = df_proc.copy()

    # ─ 회차별 야간 시간 ─
    df["야간_분"] = df.apply(
        lambda r: calc_night_minutes(r["출발분"], r["보정운행시간"]), axis=1
    )

    # ─ 일별 집계 ─
    daily = df.groupby(["운전자", "운행일_dt", "운행일", "연도", "월", "요일"]).agg(
        총시간_분=("보정운행시간", "sum"),
        야간_분=("야간_분", "sum"),
    ).reset_index()

    # ─ 공휴일 집합 ─
    pub_hol_set = set(KOREAN_HOLIDAYS.keys())

    # ─ 운전자별 지정휴일 요일 dict ─
    hol_lookup = {}
    if driver_hol_df is not None and not driver_hol_df.empty:
        hdf = driver_hol_df.copy()
        hdf.columns = [c.strip() for c in hdf.columns]
        dcol = next((c for c in ["운전자","이름","성명","name"] if c in hdf.columns), hdf.columns[0])
        hdf = hdf.rename(columns={dcol: "운전자"})
        for _, row in hdf.iterrows():
            d = str(row["운전자"]).strip()
            hol_lookup[d] = {int(row.get("지정휴일1", 5)), int(row.get("지정휴일2", 6))}

    def is_holiday(driver, dt):
        d = dt.date() if hasattr(dt, "date") else dt
        if d in pub_hol_set:
            return True
        if driver in hol_lookup and dt.weekday() in hol_lookup[driver]:
            return True
        return False

    daily["휴일여부"] = daily.apply(
        lambda r: is_holiday(r["운전자"], r["운행일_dt"]), axis=1
    )
    daily["쉬프트유형"] = daily["총시간_분"].apply(
        lambda x: "단축" if x < 300 else "정상"
    )

    # ─ 6분류 계산 ─
    def calc_cats(row):
        t     = float(row["총시간_분"])
        night = float(row["야간_분"])
        hol   = row["휴일여부"]
        shift = row["쉬프트유형"]

        if hol:
            return pd.Series({
                "A_평일8h": 0.0,
                "B_평일9h상계": 0.0,
                "C_야간": night,
                "D_쉬프트단축": 0.0,
                "E_휴일8h": min(t, 480.0),
                "F_휴일8h초과": max(t - 480.0, 0.0),
            })
        else:
            if shift == "단축":
                return pd.Series({
                    "A_평일8h": 0.0,
                    "B_평일9h상계": 0.0,
                    "C_야간": night,
                    "C_야간상계": night - 300.0,
                    "D_쉬프트단축": t - 300.0,
                    "E_휴일실제": 0.0,
                    "E_휴일8h": 0.0,
                    "F_휴일8h초과": 0.0,
                    "G_휴일상계": 0.0,
                })
            else:
                return pd.Series({
                    "A_평일8h": min(t, 480.0),
                    "B_평일9h상계": t - 540.0,
                    "C_야간": night,
                    "C_야간상계": night - 300.0,
                    "D_쉬프트단축": 0.0,
                    "E_휴일실제": 0.0,
                    "E_휴일8h": 0.0,
                    "F_휴일8h초과": 0.0,
                    "G_휴일상계": 0.0,
                })

    cats = daily.apply(calc_cats, axis=1)
    return pd.concat([daily, cats], axis=1)


def monthly_work_summary(df_cls):
    """일별 6분류 → 월별 개인별 합계 (분 및 시간)"""
    grp = df_cls.groupby(["운전자", "연도", "월"]).agg(
        근무일수=("운행일_dt", "nunique"),
        정상일수=(  "쉬프트유형", lambda x: (x=="정상").sum()),
        단축일수=(  "쉬프트유형", lambda x: (x=="단축").sum()),
        휴일일수=(  "휴일여부",   "sum"),
        A_평일8h_분=("A_평일8h", "sum"),
        B_평일9h상계_분=("B_평일9h상계", "sum"),
        C_야간_분=("C_야간", "sum"),
        C_야간상계_분=("C_야간상계", "sum"),
        D_쉬프트단축_분=("D_쉬프트단축", "sum"),
        E_휴일실제_분=("E_휴일실제", "sum"),
        E_휴일8h_분=("E_휴일8h", "sum"),
        F_휴일8h초과_분=("F_휴일8h초과", "sum"),
        G_휴일상계_분=("G_휴일상계", "sum"),
    ).reset_index()

    for col in ["A_평일8h", "B_평일9h상계", "C_야간", "C_야간상계", "D_쉬프트단축", "E_휴일실제", "E_휴일8h", "F_휴일8h초과", "G_휴일상계"]:
        grp[f"{col}_시간"] = (grp[f"{col}_분"] / 60).round(2)

    return grp


def mins_to_hhmm(mins):
    """분(float) → 'HH:MM' 문자열. 음수는 '-HH:MM'"""
    try:
        m = float(mins)
        sign = "-" if m < 0 else ""
        m = abs(m)
        h = int(m // 60)
        mi = int(round(m % 60))
        return f"{sign}{h:02d}:{mi:02d}"
    except:
        return "00:00"


def df_to_hhmm(df, cols):
    """지정 컬럼들을 분→HH:MM 으로 변환한 복사본 반환"""
    out = df.copy()
    for c in cols:
        if c in out.columns:
            out[c] = out[c].apply(mins_to_hhmm)
    return out


# ============================================================
# 시급 / 급여 계산 관련 함수
# ============================================================


@st.cache_data(show_spinner=False)
def load_payroll(_file):
    """급여명세서 xlsx → 통합 DataFrame
    컬럼: 사원번호, 성명, 급여년월, 연도, 월, 기본급, 상여금,
          연장근로, 야간오전, 야간오후, 휴일오전, 휴일오후,
          경축수당, 심야수당, 지급총액
    """
    _file.seek(0)
    xl = pd.ExcelFile(io.BytesIO(_file.read()))
    dfs = []
    for sheet in xl.sheet_names:
        try:
            df = pd.read_excel(_file, sheet_name=sheet)
            df.columns = [str(c).strip() for c in df.columns]
            if '사원번호' not in df.columns and '성명' not in df.columns:
                continue
            dfs.append(df)
        except Exception:
            continue
    if not dfs:
        return pd.DataFrame()
    out = pd.concat(dfs, ignore_index=True)
    out['사원번호'] = out['사원번호'].astype(str).str.strip().str.zfill(7)
    out['급여년월'] = out['급여년월'].astype(str).str.strip()
    out['연도'] = out['급여년월'].str[:4].astype(int)
    out['월']   = out['급여년월'].str[4:6].astype(int)
    # 지급총액 계산 (명세서 항목 합계, 년차/소급 제외)
    pay_cols = ['기본급','연장근로','야간오전','야간오후','주휴수당',
                '휴일오전','휴일오후','경축수당','심야수당','무사고포상',
                '연장1','연장2','기타수당1','기타수당2','기타수당3','교육수당']
    for c in pay_cols:
        if c not in out.columns:
            out[c] = 0
        out[c] = pd.to_numeric(out[c], errors='coerce').fillna(0)
    out['지급총액_명세'] = out[pay_cols].sum(axis=1)
    return out

@st.cache_data(show_spinner=False)
def parse_wage_table(_file):
    """시급.xlsx 통상시급(2) 시트 → {(고용형태, 호봉번호, 연도): (시급, 통상시급)}"""
    _file.seek(0)   # 파일 포인터 초기화
    wb = load_workbook(io.BytesIO(_file.read()), read_only=True, data_only=True)
    ws = wb['통상시급 (2)']
    wage_dict = {}
    current_type = '정규직'
    current_grade_num = 0

    for row in ws.iter_rows(values_only=True):
        full = list(row)
        r = [v for v in full if v is not None]
        if not r:
            continue
        col2 = full[2] if len(full) > 2 else None
        col3 = full[3] if len(full) > 3 else None
        col4 = full[4] if len(full) > 4 else None
        col5 = full[5] if len(full) > 5 else None

        if col2 and '정규직' in str(col2):
            current_type = '정규직'; current_grade_num = 0; continue
        if col2 and '촉탁직' in str(col2):
            current_type = '촉탁직'; current_grade_num = 0; continue
        if col2 and '호봉' in str(col2):
            current_grade_num += 1
        if col3 and '년도' in str(col3) and col4 and col5:
            try:
                year = int(str(col3).replace('년도', ''))
                key = (current_type, current_grade_num, year)
                wage_dict[key] = (int(col4), int(col5))
            except Exception:
                pass
    wb.close()
    # 2025년 = 2024년 값으로 대체
    for key in list(wage_dict.keys()):
        t, g, y = key
        if y == 2024:
            wage_dict[(t, g, 2025)] = wage_dict[key]
    return wage_dict


def get_tongsigeup(wage_dict, 고용형태, 호봉번호, year):
    """호봉+연도 → 통상시급 (없으면 가장 가까운 연도 반환)"""
    key = (고용형태, int(호봉번호), int(year))
    if key in wage_dict:
        return wage_dict[key][1]
    # 연도 fallback: 가장 가까운 연도
    candidates = [(t, g, y) for (t, g, y) in wage_dict if t == 고용형태 and g == int(호봉번호)]
    if not candidates:
        return 0
    closest = min(candidates, key=lambda k: abs(k[2] - int(year)))
    return wage_dict[closest][1]


def calc_wages_monthly(df_monthly, driver_info_df, wage_dict):
    """
    월별 집계 df + 운전자정보 + 시급 →
    급여 재산정액(A~F) 추가된 df 반환
    """
    df = df_monthly.copy()

    if driver_info_df is not None and not driver_info_df.empty:
        info = driver_info_df.copy()
        info.columns = [c.strip() for c in info.columns]
        # 운전자 컬럼 자동 탐지
        dcol = next((c for c in ['운전자','이름','성명'] if c in info.columns), info.columns[0])
        info = info.rename(columns={dcol: '운전자'})
        info['운전자'] = info['운전자'].astype(str).str.strip()
        # 호봉번호/고용형태 컬럼 자동 탐지 및 정규화
        for src, dst in [(['호봉번호','호봉','grade'], '호봉번호'),
                         (['고용형태','고용','type','구분'], '고용형태')]:
            found = next((c for c in src if c in info.columns), None)
            if found and found != dst:
                info = info.rename(columns={found: dst})
        if '호봉번호' not in info.columns: info['호봉번호'] = 1
        if '고용형태' not in info.columns: info['고용형태'] = '정규직'
        # 병합할 컬럼 구성 (사원번호 있으면 포함)
        merge_cols = ['운전자', '고용형태', '호봉번호']
        if '사원번호' in info.columns: merge_cols.append('사원번호')
        df = df.merge(info[merge_cols], on='운전자', how='left')
    else:
        df['고용형태'] = '정규직'
        df['호봉번호'] = 1

    df['고용형태'] = df['고용형태'].fillna('정규직')
    df['호봉번호'] = pd.to_numeric(df['호봉번호'], errors='coerce').fillna(1).astype(int)

    # 연도별 통상시급 조회
    def get_ts(row):
        return get_tongsigeup(wage_dict, row['고용형태'], row['호봉번호'], row['연도'])

    df['통상시급'] = df.apply(get_ts, axis=1)

    # 급여 재산정 (분→시간 변환 후 시급 적용)
    def wage(mins, rate):
        h = mins / 60.0
        return round(h * rate)

    ts = df['통상시급']
    df['W_A_평일8h']     = df['A_평일8h_분'].apply(lambda m: 0)                           + (df['A_평일8h_분'] / 60 * ts).round().astype(int)
    df['W_B_평일9h상계'] = (df['B_평일9h상계_분'] / 60 * ts * 1.5).round().astype(int)
    df['W_C_야간']       = (df['C_야간_분']       / 60 * ts * 0.5).round().astype(int)
    df['W_D_쉬프트단축'] = (df['D_쉬프트단축_분'] / 60 * ts * 1.5).round().astype(int)
    df['W_E_휴일8h']     = (df['E_휴일8h_분']     / 60 * ts * 1.5).round().astype(int)
    df['W_F_휴일8h초과'] = (df['F_휴일8h초과_분'] / 60 * ts * 2.0).round().astype(int)

    # 재산정 합계
    df['재산정_합계'] = (df[['W_A_평일8h','W_B_평일9h상계','W_C_야간',
                              'W_D_쉬프트단축','W_E_휴일8h','W_F_휴일8h초과']].sum(axis=1))
    return df


def merge_payroll(df_wages, driver_info_df, payroll_df):
    """재산정 df + 급여명세서 → 차액 계산"""
    if payroll_df is None or payroll_df.empty:
        return df_wages

    df = df_wages.copy()

    # 사원번호가 아직 없으면 driver_info에서 가져오기
    if '사원번호' not in df.columns:
        if driver_info_df is not None:
            info = driver_info_df.copy()
            info.columns = [c.strip() for c in info.columns]
            dcol = next((c for c in ['운전자','이름','성명'] if c in info.columns), info.columns[0])
            info = info.rename(columns={dcol: '운전자'})
            if '사원번호' in info.columns:
                info['사원번호'] = info['사원번호'].astype(str).str.strip().str.zfill(7)
                df = df.merge(info[['운전자','사원번호']], on='운전자', how='left')
            else:
                return df_wages  # 사원번호 정보 없음
        else:
            return df_wages

    # 사원번호 형식 통일
    df['사원번호'] = df['사원번호'].astype(str).str.strip().str.zfill(7)

    # 급여명세서 월별 집계
    pay_cols_avail = {
        'P_기본급':   '기본급',   'P_연장근로': '연장근로',
        'P_야간오전': '야간오전', 'P_야간오후': '야간오후',
        'P_주휴수당': '주휴수당', 'P_휴일오전': '휴일오전',
        'P_휴일오후': '휴일오후', 'P_경축수당': '경축수당',
        'P_지급총액': '지급총액_명세',
    }
    agg_dict = {k: (v, 'sum') for k, v in pay_cols_avail.items()
                if v in payroll_df.columns}
    if not agg_dict:
        return df

    pay_agg = payroll_df.groupby(['사원번호','연도','월']).agg(**agg_dict).reset_index()
    pay_agg['사원번호'] = pay_agg['사원번호'].astype(str).str.strip().str.zfill(7)

    df = df.merge(pay_agg, on=['사원번호','연도','월'], how='left')

    # 기지급 합계
    p_cols = [c for c in ['P_기본급','P_연장근로','P_야간오전','P_야간오후',
                           'P_주휴수당','P_휴일오전','P_휴일오후','P_경축수당']
              if c in df.columns]
    df['P_기지급_합계'] = df[p_cols].fillna(0).sum(axis=1)

    # 차액 = 재산정합계 - 기지급합계
    if '재산정_합계' in df.columns:
        df['차액'] = df['재산정_합계'] - df['P_기지급_합계']
    return df



def generate_report_excel(df_with_wages, df_proc_detail):
    """
    개인별 내용증명 양식 Excel 생성
    시트 구성: [전체(상세)] + [001 운전자명, 002 운전자명, ...]
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, Alignment, PatternFill,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    import io as _io

    wb = Workbook()

    # ── 헬퍼 ──
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    hdr_fill2 = PatternFill("solid", fgColor="FCE4D6")
    bold = Font(bold=True, name='맑은 고딕', size=9)
    normal = Font(name='맑은 고딕', size=9)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_al = Alignment(horizontal='right', vertical='center')
    time_fmt = '[h]:mm'
    won_fmt  = '#,##0;[Red]-#,##0'

    def set_cell(ws, row, col, value, font=None, align=None, fill=None,
                 border=None, num_format=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:      c.font = font
        if align:     c.alignment = align
        if fill:      c.fill = fill
        if border:    c.border = border
        if num_format: c.number_format = num_format
        return c

    # ── 1. 전체 상세 시트 ──
    ws_all = wb.active
    ws_all.title = '전체'

    detail_cols = ['운행일','회사명','노선','차량번호','운전자',
                   '운행출발일시','운행종료일시','보정운행시간','야간_분',
                   '이상값여부','이상값_유형']
    avail = [c for c in detail_cols if c in df_proc_detail.columns]
    hdrs  = {'운행일':'운행일','회사명':'회사명','노선':'노선',
             '차량번호':'차량번호','운전자':'운전자',
             '운행출발일시':'출발','운행종료일시':'종료',
             '보정운행시간':'운행시간(분)','야간_분':'야간(분)',
             '이상값여부':'이상값','이상값_유형':'이상값유형'}
    for ci, col in enumerate(avail, 1):
        set_cell(ws_all, 1, ci, hdrs.get(col, col), font=bold,
                 align=center, fill=hdr_fill, border=border)
    for ri, (_, row) in enumerate(df_proc_detail[avail].iterrows(), 2):
        for ci, col in enumerate(avail, 1):
            set_cell(ws_all, ri, ci, row[col], font=normal, border=border)

    # ── 2. 개인별 시트 ──
    drivers = sorted(df_with_wages['운전자'].unique())

    # 연월 범위 (2020.01 ~ 최대연도.12)
    all_years = sorted(df_with_wages['연도'].unique().astype(int))
    year_range = range(min(all_years), max(all_years) + 1)

    for idx, driver in enumerate(drivers, 1):
        drv_df = df_with_wages[df_with_wages['운전자'] == driver].copy()

        sheet_name = f"{idx:03d} {driver}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        # 통상시급 (연도별 대표값 - 해당 연도 첫 번째)
        ts_by_year = drv_df.groupby('연도')['통상시급'].first().to_dict()

        # ── 헤더 4행 ──
        # Row 1: B1:F1 병합(이름), K1(내용증명) - 병합범위 밖에 배치
        ws.merge_cells('B1:F1')
        set_cell(ws, 1, 2, f'{idx:03d}번 원고 {driver}', font=Font(bold=True, size=10, name='맑은 고딕'), align=center)
        set_cell(ws, 1, 11, '내용증명', font=bold, align=center)

        # Row 2
        ws.merge_cells('B2:I2')
        set_cell(ws, 2, 2, '1. (실제 근로시간)', font=bold, align=center, fill=hdr_fill)
        ws.merge_cells('L2:S2')
        set_cell(ws, 2, 12, '2. (실제 근로시간에 해당하는 급여 재산정액)', font=bold, align=center, fill=hdr_fill2)
        set_cell(ws, 2, 20, '통상시급', font=bold, align=center)
        hdr_fill3 = PatternFill("solid", fgColor="E2EFDA")
        ws.merge_cells('V2:AB2')
        set_cell(ws, 2, 22, '3. (실제 지급한 급여명세서)', font=bold, align=center, fill=hdr_fill3)
        set_cell(ws, 2, 29, '차액(추가지급)', font=bold, align=center)

        # Row 3
        for col, lbl in [(4,'A'),(5,'B'),(6,'C'),(7,'D'),(9,'E'),(10,'F')]:
            set_cell(ws, 3, col, lbl, font=bold, align=center, fill=hdr_fill)
        for col, lbl in [(12,'A'),(13,'B'),(14,'C'),(15,'D'),(17,'E'),(18,'F')]:
            set_cell(ws, 3, col, lbl, font=bold, align=center, fill=hdr_fill2)

        # Row 4
        h4 = {2:'년', 3:'월',
              4:'평일근무\n8시간기준', 5:'평일9시간\n월상계', 6:'야간근무', 7:'쉬프트단축',
              9:'휴일\n8시간근무', 10:'휴일\n8시간초과',
              12:'평일근무\n8시간기준', 13:'평일9시간\n월상계', 14:'야근수당',
              15:'쉬프트단축', 17:'휴일근무\n8시간', 18:'휴일근무\n8시간초과',
              20:'통상시급(원)',
              22:'기본급', 23:'연장근로', 24:'야간오전', 25:'야간오후',
              26:'휴일오전', 27:'휴일오후', 29:'차액(+추가지급)'}
        for col, lbl in h4.items():
            if col <= 10:      fill = hdr_fill
            elif col <= 18:    fill = hdr_fill2
            elif col <= 27:    fill = hdr_fill3 if 'hdr_fill3' in dir() else None
            else:              fill = None
            set_cell(ws, 4, col, lbl, font=bold, align=center, fill=fill, border=border)

        # Row 5 (배율)
        for col, lbl in [(12,'×1.0'),(13,'×1.5'),(14,'×0.5'),(15,'×1.5'),(17,'×1.5'),(18,'×2.0')]:
            set_cell(ws, 5, col, lbl, font=bold, align=center, fill=hdr_fill2, border=border)

        # ── 데이터 행 ──
        row_num = 6
        prev_year = None

        for year in year_range:
            for month in range(1, 13):
                sub = drv_df[(drv_df['연도'] == year) & (drv_df['월'] == month)]

                if sub.empty:
                    A_min = B_min = C_min = D_min = E_min = F_min = 0
                    wA = wB = wC = wD = wE = wF = 0
                    ts_val = ts_by_year.get(year, 0)
                else:
                    r = sub.iloc[0]
                    A_min = float(r.get('A_평일8h_분', 0) or 0)
                    B_min = float(r.get('B_평일9h상계_분', 0) or 0)
                    C_min = float(r.get('C_야간_분', 0) or 0)
                    D_min = float(r.get('D_쉬프트단축_분', 0) or 0)
                    E_min = float(r.get('E_휴일8h_분', 0) or 0)
                    F_min = float(r.get('F_휴일8h초과_분', 0) or 0)
                    wA = int(r.get('W_A_평일8h', 0) or 0)
                    wB = int(r.get('W_B_평일9h상계', 0) or 0)
                    wC = int(r.get('W_C_야간', 0) or 0)
                    wD = int(r.get('W_D_쉬프트단축', 0) or 0)
                    wE = int(r.get('W_E_휴일8h', 0) or 0)
                    wF = int(r.get('W_F_휴일8h초과', 0) or 0)
                    ts_val = int(r.get('통상시급', 0) or 0)

                # 연도는 첫 달에만 표시
                yr_val = year if prev_year != year else None
                set_cell(ws, row_num, 2, yr_val, font=normal, align=center, border=border)
                set_cell(ws, row_num, 3, month, font=normal, align=center, border=border)

                # 시간 컬럼 (분 → Excel 시간값 = 분/60/24)
                for col, mins in [(4, A_min),(5, B_min),(6, C_min),(7, D_min),
                                   (9, E_min),(10, F_min)]:
                    val = mins / 60.0 / 24.0  # Excel time fraction
                    c = set_cell(ws, row_num, col, val, font=normal,
                                 align=center, border=border, num_format=time_fmt)

                # 급여 컬럼
                for col, won in [(12, wA),(13, wB),(14, wC),(15, wD),(17, wE),(18, wF)]:
                    set_cell(ws, row_num, col, won, font=normal,
                             align=right_al, border=border, num_format=won_fmt)

                # 통상시급
                set_cell(ws, row_num, 20, ts_val, font=normal,
                         align=right_al, border=border, num_format='#,##0')

                # 3. 기지급 급여명세서
                if 'P_기본급' in (sub.columns if not sub.empty else []):
                    pay_row = sub.iloc[0] if not sub.empty else None
                    for col, key in [(22,'P_기본급'),(23,'P_연장근로'),
                                     (24,'P_야간오전'),(25,'P_야간오후'),
                                     (26,'P_휴일오전'),(27,'P_휴일오후')]:
                        val = int(pay_row[key]) if (pay_row is not None and key in pay_row.index and pd.notna(pay_row[key])) else 0
                        set_cell(ws, row_num, col, val, font=normal,
                                 align=right_al, border=border, num_format=won_fmt)
                    # 차액
                    diff = int(pay_row['차액']) if (pay_row is not None and '차액' in pay_row.index and pd.notna(pay_row['차액'])) else 0
                    diff_font = Font(name='맑은 고딕', size=9, color='FF0000' if diff > 0 else '000000', bold=diff != 0)
                    set_cell(ws, row_num, 29, diff, font=diff_font,
                             align=right_al, border=border, num_format=won_fmt)

                prev_year = year
                row_num += 1

        # 열 너비
        col_widths = {2:6, 3:4, 4:10, 5:10, 6:9, 7:9, 8:1, 9:9, 10:9, 11:1,
                      12:11, 13:11, 14:9, 15:9, 16:1, 17:11, 18:11, 20:10}
        for col, w in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[4].height = 36

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def to_excel(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="데이터")
    return buf.getvalue()


# ============================================================
# 사이드바 UI
# ============================================================
with st.sidebar:
    st.image("https://img.icons8.com/fluency/96/bus.png", width=60)
    st.title("운전자 근로시간 분석")
    st.caption("v1.0 · 대원교통")
    st.markdown("---")

    st.subheader("📂 파일 업로드")
    uploaded_ops = st.file_uploader(
        "① 운행데이터 (xlsx, 여러 개)",
        type=["xlsx"], accept_multiple_files=True, key="ops"
    )
    uploaded_route = st.file_uploader(
        "② 인가현황 (xls 변환본 xlsx)",
        type=["xlsx"], key="route"
    )
    uploaded_hol = st.file_uploader(
        "③ 지정휴일 (CSV)",
        type=["csv"], key="hol"
    )

    # 지정휴일 템플릿 다운로드
    tmpl = pd.DataFrame({
        "운전자": ["홍길동", "김철수", "이영희"],
        "지정휴일1": ["토", "금", "토"],
        "지정휴일2": ["일", "토", "일"],
    })
    st.download_button(
        "📥 지정휴일 템플릿",
        ("\ufeff" + tmpl.to_csv(index=False)).encode("utf-8"),
        "지정휴일_템플릿.csv", "text/csv"
    )

    st.markdown("---")
    st.subheader("⚙️ 이상값 기준")
    short_min = st.slider("최소 운행시간 (분)", 1, 30, 10,
                          help="이 값 미만이면 '미조작' 이상값으로 처리")
    extra_min = st.slider("인가시간 초과 허용 (분)", 0, 120, 60,
                          help="인가운행시간 + 이 값 초과 시 이상값 처리")
    st.caption(f"예: 2016노선(245분) → 상한 {245+extra_min}분")

    uploaded_wage = st.file_uploader(
        "④ 시급 파일 (xlsx)",
        type=["xlsx"], key="wage"
    )
    uploaded_drv_info = st.file_uploader(
        "⑤ 운전자 호봉·사번 정보 (CSV)",
        type=["csv"], key="drv_info"
    )
    uploaded_payroll = st.file_uploader(
        "⑥ 급여명세서 (xlsx, xlsb→xlsx 변환 후)",
        type=["xlsx"], key="payroll"
    )

    # 운전자 호봉 템플릿
    tmpl_drv = pd.DataFrame({
        "운전자":   ["홍길동", "김철수"],   # 운행데이터 운전자명 (정확히 일치해야 함)
        "사원번호": ["0000121", "0000300"], # 급여명세서 사원번호
        "성명":     ["홍길동", "김철수"],   # 급여명세서 성명
        "고용형태": ["정규직", "촉탁직"],
        "호봉번호": [3, 1],
    })
    st.download_button(
        "📥 운전자 호봉 템플릿",
        ("﻿" + tmpl_drv.to_csv(index=False)).encode("utf-8"),
        "운전자_호봉_템플릿.csv", "text/csv"
    )

    st.markdown("---")
    st.caption("📌 000000 미등록 운전자는 포함 처리 (추후 역추적 등록 예정)")

# ============================================================
# 메인 탭 구성
# ============================================================
tab_main, tab_anom, tab_work, tab_attend, tab_labor, tab_dl = st.tabs([
    "📊 데이터 현황",
    "⚠️ 이상값 분석",
    "⏱️ 개인별 근로시간",
    "📅 근무일수 / 만근일수",
    "⚖️ 근로시간 재산정",
    "💾 다운로드",
])

# ─── 데이터가 없을 때 안내 ───
if not uploaded_ops:
    with tab_main:
        st.info("👈 좌측 사이드바에서 운행데이터 파일을 업로드하세요.")
        st.markdown("""
        ### 📋 사용 방법
        1. **운행데이터 파일** 업로드 (월별 xlsx 파일, 여러 개 동시 가능)
        2. **인가현황 파일** 업로드 (노선별 이상값 기준 자동 설정)
        3. **지정휴일 파일** 업로드 (운전자별 만근일수 계산)
        4. 이상값 탐지 결과 확인 → 보정된 근로시간 확인 → 다운로드

        ### 🔧 처리 로직 요약
        | 항목 | 내용 |
        |------|------|
        | 근로시간 | 회차별 순수 운행시간 합산 (자정 넘김 자동 처리) |
        | 단시간 이상값 | 10분 미만 → 동일 월·요일·노선·시간대 평균으로 대체 |
        | 장시간 이상값 | 인가운행시간 + 60분 초과 → 동일 방법으로 대체 |
        | 대체 fallback | 동일 노선·월 평균 → 인가운행시간 순으로 적용 |
        | 만근일수 | 달력일수 − 지정휴일 − 비중복공휴일 |
        | 미등록(000000) | 제외 없이 포함 (추후 역추적 등록) |
        """)
    st.stop()

# ─── 데이터 로딩 ───
with st.spinner("⏳ 파일 로딩 중..."):
    df_raw = load_operation_files(uploaded_ops)

if df_raw.empty:
    st.error("운행데이터를 불러오지 못했습니다. 파일 형식을 확인하세요.")
    st.stop()

route_ref = None
if uploaded_route:
    route_ref = load_route_ref(uploaded_route)

driver_hol_df = None
if uploaded_hol:
    raw = uploaded_hol.read()
    if raw.startswith(b'\xef\xbb\xbf'):
        raw = raw[3:]
    hol = None
    for enc in ["utf-8", "cp949", "euc-kr", "latin-1"]:
        try:
            hol = pd.read_csv(io.StringIO(raw.decode(enc)))
            break
        except (UnicodeDecodeError, Exception):
            continue
    if hol is None:
        st.error("지정휴일 파일 인코딩 오류. UTF-8 또는 CP949로 저장 후 다시 업로드하세요.")
        st.stop()
    if "지정휴일1" in hol.columns:
        hol["지정휴일1"] = hol["지정휴일1"].map(WEEKDAY_MAP).fillna(5).astype(int)
    if "지정휴일2" in hol.columns:
        hol["지정휴일2"] = hol["지정휴일2"].map(WEEKDAY_MAP).fillna(6).astype(int)
    driver_hol_df = hol

# ─── 시급 / 운전자 호봉 정보 로딩 ───
# ─── 급여명세서 로딩 ───
payroll_df = None
if uploaded_payroll:
    try:
        payroll_df = load_payroll(uploaded_payroll)
        if payroll_df.empty:
            st.sidebar.warning("급여명세서 파일에서 데이터를 읽지 못했습니다.")
    except Exception as e:
        st.sidebar.warning(f"급여명세서 로드 실패: {e}")

wage_dict = {}
if uploaded_wage:
    try:
        wage_dict = parse_wage_table(uploaded_wage)
    except Exception as e:
        st.sidebar.warning(f"시급 파일 로드 실패: {e}")

driver_info_df = None
if uploaded_drv_info:
    raw_di = uploaded_drv_info.read()
    if raw_di.startswith(b'\xef\xbb\xbf'): raw_di = raw_di[3:]
    _di = None
    for enc in ["utf-8","cp949","euc-kr","latin-1"]:
        try: _di = pd.read_csv(io.StringIO(raw_di.decode(enc))); break
        except: continue
    if _di is not None:
        _di.columns = [c.strip() for c in _di.columns]
        dcol = next((c for c in ['운전자','이름','성명'] if c in _di.columns), _di.columns[0])
        _di = _di.rename(columns={dcol: '운전자'})
        _di['운전자'] = _di['운전자'].astype(str).str.strip()
        driver_info_df = _di

# ─── 처리 ───
with st.spinner("⚙️ 이상값 탐지 및 보정 중..."):
    df_proc = process_data(df_raw, route_ref, short_min, extra_min)

df_daily = build_daily_v2(df_proc)

years_list = sorted(df_daily["연도"].dropna().unique().astype(int))
months_list = sorted(df_daily["월"].dropna().unique().astype(int))

# ============================================================
# Tab 1: 데이터 현황
# ============================================================
with tab_main:
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("📄 총 레코드", f"{len(df_proc):,}건")
    c2.metric("👤 운전자 수", f"{df_proc['운전자'].nunique():,}명")
    c3.metric("🚌 노선 수", f"{df_proc['노선'].nunique():,}개")
    c4.metric("📅 시작일", str(df_proc["운행일"].min()))
    c5.metric("📅 종료일", str(df_proc["운행일"].max()))

    st.markdown("---")

    col_l, col_r = st.columns(2)
    with col_l:
        st.subheader("🗓️ 월별 운행 건수")
        monthly_cnt = df_proc.groupby(["연도", "월"]).size().reset_index(name="운행건수")
        monthly_cnt["연월"] = monthly_cnt["연도"].astype(str) + "-" + monthly_cnt["월"].astype(str).str.zfill(2)
        st.dataframe(monthly_cnt[["연월", "운행건수"]], use_container_width=True, hide_index=True)

    with col_r:
        if route_ref is not None:
            st.subheader("📋 노선별 인가운행시간 (참조)")
            st.dataframe(route_ref.sort_values(["연도", "노선"]), use_container_width=True, hide_index=True)
        else:
            st.info("인가현황 파일을 업로드하면 노선별 이상값 기준이 자동으로 설정됩니다.")

    st.markdown("---")
    st.subheader("🔍 원본 데이터 미리보기")
    disp_cols = ["운행일", "회사명", "노선", "차량번호", "운전자",
                 "운행출발일시", "운행종료일시", "운행시간(당회)"]
    st.dataframe(df_raw[disp_cols].head(200), use_container_width=True, hide_index=True)

# ============================================================
# Tab 2: 이상값 분석
# ============================================================
with tab_anom:
    anom = df_proc[df_proc["이상값여부"]].copy()
    total = len(df_proc)
    anom_cnt = len(anom)

    c1, c2, c3 = st.columns(3)
    c1.metric("⚠️ 전체 이상값", f"{anom_cnt:,}건", f"{anom_cnt/total*100:.1f}%")
    c2.metric("🔻 단시간(미조작)", f"{(anom['이상값_유형']=='단시간(미조작)').sum():,}건")
    c3.metric("🔺 장시간(과장조작)", f"{(anom['이상값_유형']=='장시간(과장조작)').sum():,}건")

    st.markdown("---")
    st.subheader("이상값 목록")

    # 필터
    cf1, cf2, cf3 = st.columns(3)
    f_year = cf1.selectbox("연도", ["전체"] + years_list, key="av_y")
    f_type = cf2.selectbox("유형", ["전체", "단시간(미조작)", "장시간(과장조작)"], key="av_t")
    f_route = cf3.selectbox("노선", ["전체"] + sorted(df_proc["노선"].unique()), key="av_r")

    filtered_anom = anom.copy()
    if f_year != "전체":
        filtered_anom = filtered_anom[filtered_anom["연도"] == int(f_year)]
    if f_type != "전체":
        filtered_anom = filtered_anom[filtered_anom["이상값_유형"] == f_type]
    if f_route != "전체":
        filtered_anom = filtered_anom[filtered_anom["노선"] == f_route]

    show_cols = ["운행일", "운전자", "노선", "운행출발일시", "운행종료일시",
                 "계산운행시간", "인가운행시간", "이상값_상한", "이상값_유형", "보정운행시간"]
    st.dataframe(
        filtered_anom[show_cols].sort_values("운행일").reset_index(drop=True),
        use_container_width=True, hide_index=True
    )

    st.caption(f"보정기준: 동일 월·요일·노선·출발시간 ±{extra_min}분 정상 평균 → 노선·월 평균 → 인가운행시간 순으로 대체")

# ============================================================
# Tab 3: 개인별 근로시간
# ============================================================
with tab_work:
    st.subheader("⏱️ 개인별 근로시간")

    c1, c2, c3 = st.columns(3)
    sel_y = c1.selectbox("연도", years_list, key="w_y")
    sel_m = c2.selectbox("월", months_list, key="w_m")
    sel_d = c3.selectbox("운전자", ["전체"] + sorted(df_daily["운전자"].unique()), key="w_d")

    filt = df_daily[(df_daily["연도"] == sel_y) & (df_daily["월"] == sel_m)]
    if sel_d != "전체":
        filt = filt[filt["운전자"] == sel_d]

    # KPI
    k1, k2, k3 = st.columns(3)
    k1.metric("운전자 수", f"{filt['운전자'].nunique():,}명")
    k2.metric("평균 일 근로시간", f"{filt['총운행시간_시간'].mean():.1f}시간")
    k3.metric("이상값 포함 일수", f"{filt['이상값포함'].sum():,}일")

    st.markdown("#### 일별 상세")
    disp = filt.rename(columns={
        "운행일": "날짜", "운전자": "운전자", "요일": "요일번호",
        "운행횟수": "회차수", "총운행시간_분": "근로시간(분)", "총운행시간_시간": "근로시간(시간)"
    })
    disp["요일"] = disp["요일번호"].map(WEEKDAY_REV)
    show = ["날짜", "운전자", "요일", "회차수", "근로시간(분)", "근로시간(시간)", "이상값포함"]
    st.dataframe(
        disp[show].sort_values(["운전자", "날짜"]).reset_index(drop=True),
        use_container_width=True, hide_index=True
    )

    st.markdown("---")
    st.markdown("#### 월별 개인별 집계")
    monthly_per = df_daily.groupby(["연도", "월", "운전자"]).agg(
        근무일수=("운행일_dt", "nunique"),
        총운행시간_분=("총운행시간_분", "sum"),
        총운행시간_시간=("총운행시간_시간", "sum"),
        이상값포함일수=("이상값포함", "sum"),
    ).reset_index()
    monthly_per["총운행시간_시간"] = monthly_per["총운행시간_시간"].round(2)
    filt2 = monthly_per[(monthly_per["연도"] == sel_y) & (monthly_per["월"] == sel_m)]
    st.dataframe(
        filt2.sort_values("총운행시간_시간", ascending=False).reset_index(drop=True),
        use_container_width=True, hide_index=True
    )

# ============================================================
# Tab 4: 근무일수 / 만근일수
# ============================================================
with tab_attend:
    st.subheader("📅 근무일수 및 만근일수")

    c1, c2 = st.columns(2)
    sel_y2 = c1.selectbox("연도", years_list, key="a_y")
    sel_m2 = c2.selectbox("월", months_list, key="a_m")

    attend = monthly_attendance(df_daily, driver_hol_df, sel_y2, sel_m2)

    _, days_in_month = calendar.monthrange(sel_y2, sel_m2)
    pub_this_month = {d: n for d, n in KOREAN_HOLIDAYS.items()
                      if d.year == sel_y2 and d.month == sel_m2}

    k1, k2, k3 = st.columns(3)
    k1.metric("📆 달력일수", f"{days_in_month}일")
    k2.metric("🎌 법정공휴일", f"{len(pub_this_month)}일")
    k3.metric("👥 운전자 수", f"{len(attend):,}명")

    if driver_hol_df is None:
        st.warning("💡 지정휴일 파일을 업로드하면 **운전자별 만근일수**가 자동 계산됩니다.")
    else:
        st.info(
            f"만근일수 = {days_in_month}(달력) − 지정휴일수 − 비중복공휴일수\n"
            f"※ 대체공휴일 포함"
        )

    st.dataframe(attend.sort_values("운전자").reset_index(drop=True),
                 use_container_width=True, hide_index=True)

    # 해당 월 공휴일 표시
    if pub_this_month:
        st.markdown("#### 해당 월 법정공휴일 목록")
        hol_list = pd.DataFrame([
            {"날짜": str(d), "요일": WEEKDAY_REV[d.weekday()], "공휴일명": n}
            for d, n in sorted(pub_this_month.items())
        ])
        st.dataframe(hol_list, use_container_width=True, hide_index=True)


# ============================================================
# Tab 5: 근로시간 재산정 (통상임금 소송용)
# ============================================================
with tab_labor:
    st.subheader("⚖️ 실제 근로시간 재산정")
    st.caption("임금협정 기준(9h/5h) 대비 실제 운행시간 기반 재산정 | 통상임금 소송용")

    with st.spinner("근로시간 6분류 계산 중..."):
        df_cls = classify_work_hours(df_proc, driver_hol_df)
        df_monthly_labor = monthly_work_summary(df_cls)

    # ── 필터 ──
    lc1, lc2, lc3 = st.columns(3)
    sel_ly = lc1.selectbox("연도", years_list, key="l_y")
    sel_lm = lc2.selectbox("월",   months_list, key="l_m")
    sel_ld = lc3.selectbox("운전자", ["전체"] + sorted(df_cls["운전자"].unique()), key="l_d")

    filt_cls = df_cls[(df_cls["연도"]==sel_ly) & (df_cls["월"]==sel_lm)]
    if sel_ld != "전체":
        filt_cls = filt_cls[filt_cls["운전자"]==sel_ld]

    filt_mo = df_monthly_labor[(df_monthly_labor["연도"]==sel_ly) & (df_monthly_labor["월"]==sel_lm)]
    if sel_ld != "전체":
        filt_mo = filt_mo[filt_mo["운전자"]==sel_ld]

    # ── KPI ──
    k1,k2,k3,k4,k5 = st.columns(5)
    k1.metric("정상쉬프트 일수", f'{filt_cls[filt_cls["쉬프트유형"]=="정상"].shape[0]:,}일')
    k2.metric("단축쉬프트 일수", f'{filt_cls[filt_cls["쉬프트유형"]=="단축"].shape[0]:,}일')
    k3.metric("휴일 근무일수",   f'{filt_cls["휴일여부"].sum():,}일')
    k4.metric("야간 총시간",     f'{filt_cls["C_야간"].sum()/60:.1f}h')
    k5.metric("9h 초과(월계)",   f'{filt_mo["B_평일9h상계_시간"].sum():.1f}h' if not filt_mo.empty else "─")

    st.markdown("---")

    # ── 일별 상세 ──
    st.markdown("#### 📋 일별 근로시간 분류 (분 단위)")
    # 음수값 강조 표시
    def highlight_neg(val):
        if isinstance(val, str) and val.startswith("-"):
            return "color: red"
        return ""

    show_daily_raw = filt_cls[[
        "운행일","운전자","휴일여부","쉬프트유형",
        "총시간_분","A_평일8h","B_평일9h상계",
        "C_야간","C_야간상계","D_쉬프트단축","E_휴일8h","F_휴일8h초과"
    ]].copy()
    # 분 → HH:MM 변환
    time_cols_d = ["총시간_분","A_평일8h","B_평일9h상계",
                   "C_야간","C_야간상계","D_쉬프트단축","E_휴일8h","F_휴일8h초과"]
    show_daily = df_to_hhmm(show_daily_raw, time_cols_d)
    show_daily.columns = [
        "날짜","운전자","휴일","쉬프트",
        "총시간","A.평일8h","B.평일9h상계",
        "C.야간","C.야간상계(기5h차감)","D.쉬프트단축","E.휴일8h","F.휴일8h초과"
    ]
    st.dataframe(
        show_daily.sort_values(["운전자","날짜"]).reset_index(drop=True)
                  .style.map(highlight_neg,
                    subset=["B.평일9h상계","C.야간상계(기5h차감)","D.쉬프트단축"]),
        use_container_width=True, hide_index=True
    )

    st.markdown("---")

    # ── 월별 집계 ──
    st.markdown("#### 📊 월별 개인별 근로시간 합계 (시간 단위)")
    st.caption("B·D 항목은 음수=기준 미달(임금 과지급 가능), 양수=기준 초과(추가 임금 발생)")

    show_mo_raw = filt_mo[[
        "운전자","근무일수","정상일수","단축일수","휴일일수",
        "A_평일8h_분","B_평일9h상계_분",
        "C_야간_분","C_야간상계_분",
        "D_쉬프트단축_분",
        "E_휴일실제_분","G_휴일상계_분"
    ]].copy()
    time_cols_m = ["A_평일8h_분","B_평일9h상계_분",
                   "C_야간_분","C_야간상계_분",
                   "D_쉬프트단축_분",
                   "E_휴일실제_분","G_휴일상계_분"]
    show_mo = df_to_hhmm(show_mo_raw, time_cols_m)
    show_mo.columns = [
        "운전자","근무일","정상일","단축일","휴일일",
        "A.평일8h(실제)",
        "B.평일9h상계",
        "C.야간(실제)",
        "C.야간상계(기5h↓)",
        "D.쉬프트단축상계",
        "E.휴일(실제)",
        "G.휴일상계",
    ]
    st.dataframe(
        show_mo.sort_values("운전자").reset_index(drop=True)
               .style.map(highlight_neg,
                 subset=["B.평일9h상계","C.야간상계(기5h↓)","D.쉬프트단축상계","G.휴일상계"]),
        use_container_width=True, hide_index=True
    )

    st.markdown("---")
    st.markdown("#### 📥 재산정 결과 다운로드")
    dl1, dl2 = st.columns(2)
    with dl1:
        st.download_button(
            "📥 일별 상세 (전체)",
            to_excel(show_daily_raw),
            f"근로시간재산정_일별_{sel_ly}{str(sel_lm).zfill(2)}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_labor_daily"
        )
    with dl2:
        st.download_button(
            "📥 월별 집계 (전체)",
            to_excel(df_monthly_labor),
            "근로시간재산정_월별전체.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_labor_monthly"
        )

    st.markdown("---")
    st.subheader("📄 내용증명 양식 Excel 생성 (개인별 시트)")

    if not wage_dict:
        st.warning("⬅️ 사이드바에서 **④ 시급 파일**을 업로드하면 급여 재산정액까지 포함된 내용증명 Excel을 생성할 수 있습니다.")
    else:
        if driver_info_df is None:
            st.info("💡 **⑤ 운전자 호봉 정보** 파일이 없으면 전원 정규직 1호봉으로 계산됩니다.")

        # 필터 옵션
        ec1, ec2 = st.columns(2)
        sel_report_drivers = ec1.multiselect(
            "생성할 운전자 선택 (비워두면 전원)",
            options=sorted(df_monthly_labor["운전자"].unique()),
            key="report_drivers"
        )

        if st.button("📄 내용증명 Excel 생성", type="primary", key="gen_report"):
            with st.spinner("Excel 생성 중... (운전자 수에 따라 1~2분 소요)"):
                # 급여 계산
                df_wages = calc_wages_monthly(df_monthly_labor, driver_info_df, wage_dict)

                # 급여명세서 차액 합산
                if payroll_df is not None:
                    df_wages = merge_payroll(df_wages, driver_info_df, payroll_df)

                # 운전자 필터
                if sel_report_drivers:
                    df_wages = df_wages[df_wages["운전자"].isin(sel_report_drivers)]

                # Excel 생성
                excel_bytes = generate_report_excel(df_wages, df_proc)

            st.success(f"✅ {df_wages['운전자'].nunique():,}명 × 연도별 월간 집계 완료!")
            st.download_button(
                "💾 내용증명 Excel 다운로드",
                excel_bytes,
                "내용증명_근로시간재산정.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_report"
            )

    if driver_hol_df is None:
        st.info("💡 지정휴일 파일을 업로드하면 운전자별 지정휴일이 '휴일'로 정확히 분류됩니다.")


# ============================================================
# Tab 6: 다운로드
# ============================================================
with tab_dl:
    st.subheader("💾 결과 다운로드")

    st.markdown("#### 선택 다운로드")
    dc1, dc2, dc3 = st.columns(3)

    # ① 처리된 운행 데이터
    proc_cols = ["운행일", "회사명", "노선", "차량번호", "운전자",
                 "운행출발일시", "운행종료일시",
                 "계산운행시간", "인가운행시간", "이상값여부", "이상값_유형", "보정운행시간"]
    with dc1:
        st.markdown("**① 처리된 운행 데이터**")
        st.caption("이상값 탐지·보정 결과 포함")
        st.download_button(
            "📥 다운로드",
            to_excel(df_proc[proc_cols]),
            "처리된_운행데이터.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_proc"
        )

    # ② 개인별 일별 근로시간
    with dc2:
        st.markdown("**② 개인별 일별 근로시간**")
        st.caption("운전자 × 날짜 기준 합산")
        st.download_button(
            "📥 다운로드",
            to_excel(df_daily),
            "개인별_일별_근로시간.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_daily"
        )

    # ③ 월별 개인별 집계
    monthly_all = df_daily.groupby(["연도", "월", "운전자"]).agg(
        근무일수=("운행일_dt", "nunique"),
        총운행시간_분=("총운행시간_분", "sum"),
        총운행시간_시간=("총운행시간_시간", "sum"),
    ).reset_index()
    monthly_all["총운행시간_시간"] = monthly_all["총운행시간_시간"].round(2)
    with dc3:
        st.markdown("**③ 월별 개인별 집계**")
        st.caption("근무일수 + 총 근로시간")
        st.download_button(
            "📥 다운로드",
            to_excel(monthly_all),
            "월별_개인별_집계.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_monthly"
        )

    st.markdown("---")
    st.markdown("#### 이상값 전체 목록")
    st.download_button(
        "📥 이상값 목록 다운로드",
        to_excel(df_proc[df_proc["이상값여부"]][proc_cols]),
        "이상값_목록.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_anom"
    )
